import sys
from openpyxl import load_workbook

#This script reads the binaray matrix for a phylogenetic tree from an excel file and converts it into newick format.
#The first argument specifys the input directory. The second and third argument is optional and define the output directory
#and the writing mode (use "w" in the third argument to overwrite and "a" to append to existing file).

#Arguments: inputDir <outputDir> <mode>

wb = load_workbook(sys.argv[1])
ws = wb.active

outputDir = 'newick output.txt'
mode = 'w'
IndexStart = 0
cLength = 0
rLength = 0

if len(sys.argv) > 2:
    outputDir = sys.argv[2]
if len(sys.argv) > 3:
    mode = sys.argv[3]

for row in ws.columns[0]:
    cLength += 1

for col in ws.rows[0]:
    rLength += 1

subTrees = ['e' for stuff in range(cLength)]


def create_clade(inp):
    result = '(' + inp[0]
    for element in inp[1:]:
        if element == '0':
            break
        result = result + ',' + element
    return result + ')'


i = 0
while i < cLength-2:
    buffer = ['0' for data in range(cLength)]
    bufferIndex = 0
    check = True
    j = 0
    while j < rLength+1:
        if ws.cell(row=j+1, column=i+2).value == 1:
            if check:
                IndexStart = j
                check = False
            if subTrees[j] != 'e' and subTrees[j] != 'x':
                buffer[bufferIndex] = subTrees[j]
                bufferIndex += 1
                subTrees[j] = 'x'
            elif subTrees[j] == 'e':
                valueString = ws.cell(row=j+1, column=1).value
                buffer[bufferIndex] = valueString.replace(' ', '_')
                bufferIndex += 1
                subTrees[j] = 'x'
        j += 1
    subTrees[IndexStart] = create_clade(buffer)
    i += 1
subTrees[0] += ';'
print(subTrees[0])

fobj_out = open(outputDir, mode)
fobj_out.write(subTrees[0] + '\n')
fobj_out.close()
